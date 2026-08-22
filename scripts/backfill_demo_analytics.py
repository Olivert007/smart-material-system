# -*- coding: utf-8 -*-
"""补齐演示库趋势分析缺数据板块（真实数据优先，幂等）。

背景
----
build_demo_env.py 只发布库存与流水两个域，样例台账没有区域/最低库存列，
也没有资产清查数据，导致趋势分析里以下图表为空：
  库存区域分布、低库存 TOP（库存 < 最低库存）、资产公司分布、
  资产区域（域）分布、资产购买年份分布

真实数据来源（默认取仓库父目录下 /workspace/2026-07 的真实 Excel，
可用 REAL_DATA_DIR 环境变量覆盖）：
1. fact_asset：2026年通信部成都分部资产清查汇总表（成、溪、向）初稿(1).xlsx
   —— 公司 / 域 / 购买日期等均为真实值；文件缺失时退化为合成演示行。
2. fact_inventory.min_qty：副本 通信部成都分部…物资台账.XLSX 中
   「维护材料」表的「最低库存阈值」列，按物资名称匹配真实阈值；
   未匹配到的行再用合成兜底。
3. fact_inventory.region：由真实 location（存放位置）关键字推导。

口径：真实数据优先，只有缺失/不可用才用合成兜底，保证图表始终有数；
回填行带 row_key 标记（asset-real-* / asset-demo-*），可随时区分与清理。

用法
----
    python3 scripts/backfill_demo_analytics.py [material.duckdb 路径，默认 demo_data/runtime/material.duckdb]
"""
from __future__ import annotations

import hashlib
import os
import re
import sys
from datetime import datetime
from pathlib import Path

import duckdb
import openpyxl

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DB = ROOT / "demo_data" / "runtime" / "material.duckdb"
# 真实数据默认目录：仓库父目录（本项目为 /workspace/2026-07）
DEFAULT_REAL_DIR = Path(os.environ.get("REAL_DATA_DIR", "") or ROOT.parent)

ASSET_XLSX = "2026年通信部成都分部资产清查汇总表（成、溪、向）初稿(1).xlsx"
LEDGER_XLSX = (
    "副本 通信部成都分部工器具、低值易耗品、备品备件、维护材料、"
    "应急备汛物资台账 (305 - CYPC-305-002362-2026 - 1 - B) - 1 - 副本(1).XLSX"
)


def _h(text: str, salt: str) -> int:
    """稳定哈希，保证同一物资每次得到相同结果（幂等）。"""
    return int(hashlib.md5(f"{text}{salt}".encode("utf-8")).hexdigest(), 16)


def _backfill_region(con: duckdb.DuckDBPyConnection) -> int:
    """从存放位置推导区域（真实数据）：溪洛渡 / 向家坝 / 成都 / 其他。"""
    cur = con.execute(
        """
        UPDATE fact_inventory SET region = CASE
          WHEN location LIKE '%溪洛渡%' OR location LIKE '%溪右%'
               OR location LIKE '%溪左%' OR location LIKE '%溪控%' THEN '溪洛渡'
          WHEN location LIKE '%向家坝%' THEN '向家坝'
          WHEN location LIKE '%成都%' OR location LIKE '%三峡%' THEN '成都'
          ELSE '其他'
        END
        """
    )
    row = cur.fetchone()
    return int(row[0]) if row else 0


# ---------- 最低库存：真实阈值优先 ----------


def _load_real_min_qty(real_dir: Path) -> dict[str, float]:
    """从「维护材料」表读取 名称 → 最低库存阈值（真实数据）。"""
    path = real_dir / LEDGER_XLSX
    if not path.exists():
        return {}
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    if "维护材料" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["维护材料"]
    it = ws.iter_rows(values_only=True)
    next(it)  # 标题行
    next(it)  # 注意行
    header = next(it)
    name_idx = next((i for i, h in enumerate(header) if h and "名称" in str(h)), 1)
    min_idx = next((i for i, h in enumerate(header) if h and "最低库存" in str(h)), 12)
    out: dict[str, float] = {}
    for row in it:
        nm = str(row[name_idx] or "").strip()
        if not nm:
            continue
        try:
            v = float(row[min_idx])
        except (TypeError, ValueError):
            continue
        if nm not in out:  # 同名多行取首个非空阈值
            out[nm] = v
    wb.close()
    return out


def _synth_min_qty(con: duckdb.DuckDBPyConnection) -> int:
    """合成兜底：无库存行必低库存，其余按哈希制造约 1/8 的低库存项。"""
    rows = con.execute(
        "SELECT inventory_id, material_id, COALESCE(stock_qty, 0) "
        "FROM fact_inventory WHERE min_qty IS NULL"
    ).fetchall()
    n = 0
    for inventory_id, material_id, stock in rows:
        h1 = _h(str(material_id), "min-qty")
        if stock <= 0:
            min_qty = round(5 + (h1 % 40) / 2, 2)
        elif h1 % 8 == 0:
            min_qty = round(stock * (1.1 + (h1 % 50) / 100), 2)
        else:
            min_qty = round(stock * (0.4 + (h1 % 40) / 100), 2)
        con.execute(
            "UPDATE fact_inventory SET min_qty = ? WHERE inventory_id = ?",
            [min_qty, inventory_id],
        )
        n += 1
    return n


def _backfill_min_qty(
    con: duckdb.DuckDBPyConnection, real_min: dict[str, float]
) -> tuple[int, int]:
    """最低库存：命中名称的行用真实阈值覆盖，其余合成兜底。"""
    real_n = 0
    for name, threshold in real_min.items():
        cur = con.execute(
            "UPDATE fact_inventory SET min_qty = ? "
            "WHERE inventory_id IN ("
            "  SELECT i.inventory_id FROM fact_inventory i "
            "  JOIN dim_material d USING (material_id) "
            "  WHERE d.material_name = ?"
            ")",
            [threshold, name],
        )
        row = cur.fetchone()
        real_n += int(row[0]) if row else 0
    synth_n = _synth_min_qty(con)
    return real_n, synth_n


# ---------- 资产清查：真实 Excel 优先 ----------


def _norm_purchase_date(v: object) -> str | None:
    """购买日期归一化为 YYYY-MM-DD；解析失败返回 None（图表将跳过该行）。"""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v or "").strip()
    for sep in ("/", "-", "."):
        if s.count(sep) >= 2:
            try:
                y, m, d = (int(x) for x in s.split(sep)[:3])
                return f"{y:04d}-{m:02d}-{d:02d}"
            except ValueError:
                pass
    return s if re.match(r"^\d{4}", s) else None


def _sheet_status_map(wb: openpyxl.Workbook, keyword: str, status: str) -> dict[str, str]:
    """按 sheet 名关键字收集 实物资产编码 → 状态（及核对结果）。"""
    out: dict[str, str] = {}
    for ws in wb.worksheets:
        if keyword not in ws.title:
            continue
        it = ws.iter_rows(values_only=True)
        header = next(it)
        code_idx = next((i for i, h in enumerate(header) if h and "实物资产编码" in str(h)), 2)
        result_idx = next(
            (i for i, h in enumerate(header) if h and "核对结果" in str(h)), None
        )
        for row in it:
            code = str(row[code_idx] or "").strip()
            if not code or code in out:
                continue
            note = str(row[result_idx] or "").strip() if result_idx is not None else ""
            out[code] = f"{status}|{note}"
    return out


def _load_real_asset_rows(real_dir: Path) -> list[dict]:
    """读取真实资产清查汇总表：资产总表为基础，按 有问题/待报废 表补状态。"""
    path = real_dir / ASSET_XLSX
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    try:
        status_map = _sheet_status_map(wb, "有问题", "有问题")
        status_map.update(_sheet_status_map(wb, "待报废", "待报废"))
        ws = next((s for s in wb.worksheets if "资产总表" in s.title), None)
        if ws is None:
            return []
        it = ws.iter_rows(values_only=True)
        header = next(it)
        idx = {
            "company": next((i for i, h in enumerate(header) if h and "公司" == str(h).strip()), 0),
            "domain": next((i for i, h in enumerate(header) if h and str(h).strip() == "域"), 1),
            "code": next((i for i, h in enumerate(header) if h and "实物资产编码" in str(h)), 2),
            "name": next((i for i, h in enumerate(header) if h and "实物资产名称" in str(h)), 3),
            "user": next((i for i, h in enumerate(header) if h and "使用人姓名" in str(h)), 4),
            "location": next((i for i, h in enumerate(header) if h and "位置描述" in str(h)), 5),
            "date": next((i for i, h in enumerate(header) if h and "购买日期" in str(h)), 8),
            "manager": next((i for i, h in enumerate(header) if h and "管理者姓名" in str(h)), 11),
        }
        rows: list[dict] = []
        seen: set[str] = set()
        for row in it:
            code = str(row[idx["code"]] or "").strip()
            if not code or code in seen:
                continue
            seen.add(code)
            status, check = (status_map.get(code) or "正常|").split("|", 1)
            rows.append(
                {
                    "asset_code": code,
                    "asset_name": str(row[idx["name"]] or "").strip() or code,
                    "company": str(row[idx["company"]] or "").strip() or None,
                    "domain": str(row[idx["domain"]] or "").strip() or None,
                    "user_name": str(row[idx["user"]] or "").strip() or None,
                    "location": str(row[idx["location"]] or "").strip() or None,
                    "purchase_date": _norm_purchase_date(row[idx["date"]]),
                    "manager": str(row[idx["manager"]] or "").strip() or None,
                    "status": status,
                    "check_result": check or None,
                    "source_sheet": ws.title,
                }
            )
        return rows
    finally:
        wb.close()


_COMPANIES = ["CTGCY"] * 60 + ["CYPC"] * 25 + ["CTGYC"] * 15
_DOMAINS = ["TDCD"] * 45 + ["TD"] * 25 + ["TDKD"] * 20 + ["GZB"] * 10
_YEARS = [
    (2015, 2), (2016, 3), (2017, 3), (2018, 4), (2019, 8), (2020, 12),
    (2021, 12), (2022, 25), (2023, 15), (2024, 10), (2025, 6),
]
_YEAR_TOTAL = sum(w for _, w in _YEARS)
_REGION_LOCATION = {"TDCD": "成都三峡大厦", "TD": "向家坝", "TDKD": "溪洛渡", "GZB": "葛洲坝"}
_ASSET_COLUMNS = (
    "asset_code, asset_name, row_key, company, domain, user_name, manager, "
    "location, purchase_date, status, check_result, material_code, asset_qty, "
    "unit, is_instrument, replace_cycle, check_cycle, consumption_plan, "
    "tool_source, asset_quota_qty, remark, source_file, color_flag, "
    "source_sheet, source_release_id"
)


def _pick_year(v: int) -> int:
    r = v % _YEAR_TOTAL
    for year, w in _YEARS:
        if r < w:
            return year
        r -= w
    return _YEARS[-1][0]


def _synth_asset(con: duckdb.DuckDBPyConnection) -> int:
    """合成兜底：基于 dim_material 生成资产演示行（真实 Excel 缺失时）。"""
    mats = con.execute(
        "SELECT d.material_id, d.material_code, d.material_name, d.spec, d.unit "
        "FROM dim_material d "
        "WHERE d.material_id IN (SELECT DISTINCT material_id FROM fact_inventory) "
        "ORDER BY d.material_id LIMIT 120"
    ).fetchall()
    n = 0
    for i, (material_id, material_code, material_name, spec, unit) in enumerate(mats):
        h1 = _h(str(material_id), "asset")
        year = _pick_year(h1)
        code = f"ZC-{year}-{i + 1:04d}"
        company = _COMPANIES[h1 % len(_COMPANIES)]
        domain = _DOMAINS[(h1 >> 4) % len(_DOMAINS)]
        name = f"{material_name or '未命名物资'}{'·' + spec if spec else ''}"
        status = "正常" if h1 % 10 else "有问题"
        month = 1 + h1 % 12
        day = 1 + (h1 >> 8) % 28
        con.execute(
            f"INSERT INTO fact_asset ({_ASSET_COLUMNS}) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            [
                code, name, f"asset-demo-{i + 1}", company, domain,
                None, None, _REGION_LOCATION.get(domain, "成都"),
                f"{year}-{month:02d}-{day:02d} 00:00:00", status,
                "账实相符" if h1 % 10 else "账实不符", material_code,
                float(1 + h1 % 5), unit or "个", None, None, None, None, None,
                None, None, "通信部成都分部资产清查汇总表（演示样例）.xlsx",
                None, "资产清查（演示合成）", None,
            ],
        )
        n += 1
    return n


def _insert_asset_rows(con: duckdb.DuckDBPyConnection, rows: list[dict], prefix: str) -> int:
    for i, r in enumerate(rows):
        con.execute(
            f"INSERT INTO fact_asset ({_ASSET_COLUMNS}) "
            "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
            [
                r["asset_code"], r["asset_name"], f"{prefix}-{i + 1}",
                r.get("company"), r.get("domain"), r.get("user_name"),
                r.get("manager"), r.get("location"), r.get("purchase_date"),
                r.get("status"), r.get("check_result"), None, None, None,
                None, None, None, None, None, None, None,
                ASSET_XLSX, None, r.get("source_sheet"), None,
            ],
        )
    return len(rows)


def _backfill_asset(
    con: duckdb.DuckDBPyConnection, real_dir: Path
) -> tuple[int, str]:
    """资产清查：真实 Excel 优先，缺失时合成兜底。"""
    con.execute(
        "DELETE FROM fact_asset WHERE row_key LIKE 'asset-demo-%' OR row_key LIKE 'asset-real-%'"
    )
    real_rows = _load_real_asset_rows(real_dir)
    if real_rows:
        return _insert_asset_rows(con, real_rows, "asset-real"), "real"
    return _synth_asset(con), "synth"


def backfill_demo_analytics(
    db_path: str | Path, real_data_dir: str | Path | None = None
) -> dict[str, object]:
    """对指定演示库执行补齐，返回各板块影响统计。"""
    db = Path(db_path)
    if not db.exists():
        raise FileNotFoundError(f"演示库不存在: {db}")
    real_dir = Path(real_data_dir) if real_data_dir else DEFAULT_REAL_DIR
    real_min = _load_real_min_qty(real_dir)
    con = duckdb.connect(str(db))
    try:
        region_n = _backfill_region(con)
        min_real, min_synth = _backfill_min_qty(con, real_min)
        asset_n, asset_src = _backfill_asset(con, real_dir)
    finally:
        con.close()
    return {
        "region": region_n,
        "min_qty": {"real": min_real, "synth": min_synth},
        "asset": {"rows": asset_n, "source": asset_src},
    }


def main() -> int:
    db = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_DB
    print("backfill db:", db)
    print("real data dir:", DEFAULT_REAL_DIR)
    summary = backfill_demo_analytics(db)
    print("backfill summary:", summary)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
