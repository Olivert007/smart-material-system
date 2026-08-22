# -*- coding: utf-8 -*-
"""从本地原件生成脱敏演示样例（原件不入库）。

用法:
  export RAW_SAMPLE=/path/to/source-ledger.xlsx
  python3 scripts/build_desensitized_sample.py

输出: demo_data/samples/desensitized-sample.xlsx
"""
from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "demo_data" / "samples"
OUT = OUT_DIR / "desensitized-sample.xlsx"
_raw = os.environ.get("RAW_SAMPLE", "").strip()
if not _raw:
    raise SystemExit("请设置 RAW_SAMPLE=/path/to/source-ledger.xlsx（原件不入库）")
SRC = Path(_raw)

# ---------- 站点名脱敏映射（长串优先） ----------
SITE_MAP = [
    ("成都三峡大厦负一楼材料室", "望湖大厦负一楼材料室"),
    ("成都三峡大厦负1楼材料室", "望湖大厦负1楼材料室"),
    ("三峡梯调", "云岭梯调"),
    ("三峡E购", "云岭E购"),
    ("带回成都", "带回演示区域"),
    ("带至成都", "带至演示区域"),
    ("成都带来", "演示区域带来"),
    ("成都望湖大厦", "望湖大厦"),
    ("成都区域", "演示区域"),
    ("成都", "演示区域"),
    ("溪洛渡213材料室", "云岭电站213材料室"),
    ("溪洛渡设代楼316仓库", "云岭电站设代楼316仓库"),
    ("溪洛渡212/503办公室", "云岭电站212/503办公室"),
    ("溪洛渡212办公室", "云岭电站212办公室"),
    ("溪洛渡503办公室", "云岭电站503办公室"),
    ("溪左通信机房", "云岭电站左岸通信机房"),
    ("溪控通信机房", "云岭电站中控通信机房"),
    ("溪右通信/电源机房", "云岭电站右岸通信/电源机房"),
    ("向家坝左岸14号楼2层215室", "锦澜电站左岸14号楼2层215室"),
    ("向家坝14号楼215", "锦澜电站14号楼215"),
    ("向家坝集控楼304", "锦澜电站集控楼304"),
    ("成都三峡大厦六楼通信材料室", "望湖大厦6楼通信材料室"),
    ("成都三峡大厦6楼通信材料室", "望湖大厦6楼通信材料室"),
    ("成都三峡大厦负1楼材料室", "望湖大厦负1楼材料室"),
    ("成都三峡大厦6楼仓库", "望湖大厦6楼仓库"),
    ("向家坝", "锦澜电站"),
    ("溪洛渡", "云岭电站"),
    ("三峡大厦", "望湖大厦"),
    ("成都区域", "演示区域"),
    # 流水/备注文本中的简称残留（2026-08-19 补漏）
    ("溪控", "云岭中控"),
    ("溪建", "云岭建管"),
    ("宜昌", "青江"),
]

# ---------- 真实人名 → 虚构人名（长名优先替换） ----------
NAME_MAP = [
    ("张停伟", "张伟"), ("李昭源", "李敏"), ("李朝源", "李敏"),
    ("鲁睿林", "鲁铭"), ("刘继源", "刘杰"), ("毛友兵", "毛强"),
    ("赵梦柯", "赵磊"), ("冯森林", "冯浩"), ("刘成春", "刘涛"),
    ("强亚倩", "强莹"), ("赵明贵", "赵勇"), ("吕志超", "吕鑫"),
    ("周翼虎", "周亮"), ("王丹翔", "王琪"), ("夏万昌", "夏峰"),
    ("曾泽阳", "曾凯"), ("张博瑞", "张锐"), ("沈鸿", "沈立"),
    ("鲁睿林", "鲁铭"), ("杨鑫", "杨凯"), ("徐吉", "徐刚"),
    ("吕玮智", "吕川"), ("詹雯", "詹雪"), ("潘宇", "潘峰"),
    ("卓莹", "卓然"), ("李庚", "李昊"), ("梁稣", "梁平"),
]


def mask(text: str) -> str:
    t = str(text)
    for k, v in SITE_MAP:
        t = t.replace(k, v)
    for k, v in NAME_MAP:
        t = t.replace(k, v)
    return t


def main() -> int:
    wb = load_workbook(SRC, data_only=True)
    ws = wb.active

    # 数据行：跳过标题(R1)/表头(R2)/示例行(R3)，空行剔除；取 19 列
    rows = []
    for r in range(4, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, 20)]
        if v[0] is None and all(x in (None, "") for x in v[1:]):
            continue
        rows.append(v)

    # 表头：R2 的多行文本合并为单行
    headers = []
    for c in range(1, 20):
        h = ws.cell(2, c).value
        if isinstance(h, str):
            h = h.replace("\n", "").strip()
        headers.append(h or "")

    # 写新文件（干净模板，仅 19 列）
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "演示区域ZW物资台账（新模板）"
    ws2.append(["演示区域ZW物资汇总表（脱敏样例·通信物资）"])
    ws2.append(headers)
    seq = 0
    for v in rows:
        masked = [mask(x) if x is not None else None for x in v]
        if masked[0] is None or masked[0] == "":
            seq += 1
            masked[0] = seq
        else:
            try:
                seq = int(masked[0])
            except (TypeError, ValueError):
                seq += 1
                masked[0] = seq
        ws2.append(masked)

    wb2.save(str(OUT))

    # 脱敏自检：全文件扫描敏感词
    wb3 = load_workbook(OUT, data_only=True)
    ws3 = wb3.active
    alltext = "\n".join(
        str(c.value) for row in ws3.iter_rows() for c in row if c.value is not None
    )
    forbidden = ["溪洛渡", "向家坝", "三峡", "成都", "西坝", "溪控", "溪建", "宜昌",
                 "张停伟", "沈鸿", "鲁睿林", "杨鑫", "赵梦柯", "吕玮智", "李昭源", "李朝源",
                 "毛友兵", "刘继源", "冯森林", "刘成春", "强亚倩", "赵明贵", "吕志超",
                 "周翼虎", "王丹翔", "夏万昌", "曾泽阳", "张博瑞", "潘宇", "卓莹", "李庚",
                 "梁稣", "徐吉"]
    hits = [k for k in forbidden if k in alltext]
    print("脱敏输出:", OUT)
    print("数据行数:", len(rows), "| 表头列数:", len(headers))
    if hits:
        print("!! 脱敏自检未通过，仍有敏感词:", hits)
        return 1
    print("脱敏自检通过：未发现敏感词。")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
