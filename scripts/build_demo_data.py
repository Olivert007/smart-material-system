# -*- coding: utf-8 -*-
"""生成参赛演示用脱敏台账（虚构站点/人员/物料，保留真实数据结构特征）。

设计要点：
- sheet 名与列结构完全对齐 data/flow_config/ledger_route.json 与 305b_*.json，
  保证路由/流水解析/映射正常命中（维护材料→inventory+flow，备品备件→inventory+flow，
  应急备汛物资→inventory，公用工器具→asset）。
- 所有站点、人员、项目、品牌、编码均为虚构，无任何真实内部数据。
- 保留真实台账的脏数据特征：序号「例」示例行、空字段、口径不一致（现有库存≠初始+入库-出库）、
  出库记录含时间/人员/用途的混合格式文本，用于展示勾稽差异与流水智能拆解能力。

用法: python3 scripts/build_demo_data.py [输出 xlsx 路径]
默认输出: demo_data/samples/synthetic-sample.xlsx
"""
from __future__ import annotations

import random
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent

# ---------------------------------------------------------------------------
# 虚构字典（全部为虚构值，避免任何真实站点/人名/项目/品牌）
# ---------------------------------------------------------------------------
SITES = ["青岩水电站", "云岭电站", "锦澜电站", "白鹭调控中心", "望湖大厦"]
LOCS = [
    "青岩水电站东区仓库",
    "青岩水电站调度大楼",
    "云岭电站右岸中控室",
    "云岭电站大坝仓库",
    "锦澜电站左岸厂房仓库",
    "白鹭调控中心机房",
    "望湖大厦6楼通信材料室",
    "望湖大厦负一楼材料室",
]
PEOPLE = ["张伟", "李娜", "王强", "陈静", "刘洋", "赵磊", "周敏", "孙浩"]
PROJECTS = ["2025年度备品备件补充采购", "设备维护专项采购", "年度防汛物资补充采购", "机房改造专项采购"]
SOURCES = ["电商平台采购", "统一集中采购", "年度框架协议采购"]
SYSTEMS = ["电源系统", "通信系统", "监控系统", "消防系统"]
# 通用物料池（只保留通用品名，不含品牌型号等可识别信息）
WEIHU_MATS = [
    ("插线板", "通用插线板 3米", "个"),
    ("熔纤盘", "标准熔纤盘 12芯", "个"),
    ("光模块", "千兆光模块 SFP", "个"),
    ("网线", "超五类网线", "米"),
    ("尾纤", "标准尾纤 3米", "根"),
    ("法兰盘", "LC法兰盘", "个"),
    ("扎带", "尼龙扎带 300mm", "包"),
    ("电池", "12V密封铅酸电池", "个"),
    ("整流模块", "标准整流模块 48V/40A", "个"),
    ("防雷模块", "三级防雷模块", "个"),
    ("电话机", "调度电话机", "台"),
    ("对讲机", "数字对讲机", "台"),
]
BEIPIN_MATS = [
    ("整流模块", "标准整流模块 48V/40A", "个"),
    ("风扇模块", "机柜散热风扇模块", "个"),
    ("监控硬盘", "监控级硬盘 4TB", "块"),
    ("采集单元", "数据采集单元", "个"),
    ("蓄电池", "阀控式蓄电池 12V/100Ah", "只"),
    ("断路器", "小型断路器 2P/32A", "个"),
    ("接触器", "交流接触器 40A", "个"),
    ("光纤跳线", "LC-LC光纤跳线 5米", "根"),
    ("电源模块", "直流电源模块", "个"),
    ("控制板卡", "主控板卡", "块"),
]
YINGJI_MATS = [
    ("卫星通信终端", "标准型卫星通信终端", "台"),
    ("应急照明灯", "移动应急照明灯", "盏"),
    ("潜水泵", "便携式潜水泵", "台"),
    ("救生衣", "成人救生衣", "件"),
    ("编织袋", "防汛编织袋", "条"),
    ("铁锹", "防汛铁锹", "把"),
    ("沙袋", "防汛沙袋", "条"),
    ("应急电源", "移动应急电源 3kW", "台"),
]
GONGJU_MATS = [
    ("水平尺", "铝制水平尺 600mm", "个"),
    ("活动扳手", "活动扳手 8寸", "把"),
    ("斜口钳", "斜口钳 5寸", "把"),
    ("万用表", "数字万用表", "台"),
    ("绝缘手套", "绝缘手套 10kV", "双"),
    ("安全帽", "工程安全帽", "顶"),
    ("手电筒", "强光手电筒", "个"),
    ("电钻", "手电钻 12V", "台"),
    ("卷尺", "卷尺 5米", "把"),
    ("热成像仪", "手持热成像仪", "台"),
    ("标签打印机", "便携标签打印机", "台"),
    ("网线钳", "网络压线钳", "把"),
]

WEIHU_HEADERS = [
    "序号", "名称", "品牌型号规格", "现有库存\n(=初始+入库-出库)", "单位",
    "存放位置", "保管人", "初始库存",
    "入库记录（入库时间、经手人、物资编码）", "入库数量",
    "出库记录（含时间、领用人、用途）", "出库数量",
    "最低库存阈值", "备注",
]
BEIPIN_HEADERS = [
    "序号", "物资名称", "品牌/规格型号", "存放位置", "保管人", "现有库存量",
    "初始库存", "单位", "入库记录（入库时间、入库来源、物资编码）", "入库数量",
    "出库记录（出库时间、出库原因）", "出库数量", "所属系统", "项目名称",
    "消耗计划", "物资来源", "定额数量", "公司仓库数量", "备注",
]
YINGJI_HEADERS = [
    "区域", "序号", "物资类别", "物资编码", "新集团编码", "物资名称", "型号规格",
    "计量单位", "定额数量", "现有数量", "存放货位", "管理员", "备注",
    "是否框架物资", "协议供应商名称", "推荐框架物资编码", "推荐框架物资名称",
    "推荐框架物资型号规格", "推荐框架物资供应商", "应急供应商名称",
]
GONGJU_HEADERS = [
    "序号", "物资名称", "规格型号", "物资编码", "资产编码", "数量", "单位",
    "存放位置", "保管人", "是否仪器仪表", "更换周期（年）", "检测周期（年）",
    "消耗计划", "购买日期", "工器具来源", "定额数量", "备注",
]


def fmt_in_text(i: int, rng: random.Random) -> str:
    """入库记录：时间/经手人/物资编码（混合格式，供流水拆解展示）。"""
    kinds = [
        f"2025-0{(i % 9) + 1}-1{(i % 7) + 2} 00:00:00/{rng.choice(PEOPLE)}/DM2026-{i:04d}",
        f"2025.{(i % 8) + 2}.{(i % 20) + 1}/{rng.choice(PEOPLE)}/统一采购",
        f"2026年{(i % 6) + 1}月/{rng.choice(PEOPLE)}/DM2026-{i:04d}",
        f"2026/{(i % 7) + 2}/{i % 19 + 1}/{rng.choice(PEOPLE)}/年度采购",
    ]
    return rng.choice(kinds)


def fmt_out_text(i: int, rng: random.Random) -> str:
    """出库记录：时间/领用人/用途（混合格式）。"""
    uses = ["现场检修使用", "调度大厅备用", "机房改造施工", "防汛演练消耗", "设备更换使用"]
    kinds = [
        f"2026.{i % 6 + 1}.{(i % 20) + 1}/{rng.choice(PEOPLE)}/{rng.choice(uses)}",
        f"2026-0{i % 6 + 1}-{(i % 20) + 1} 00:00:00/{rng.choice(PEOPLE)}/{rng.choice(uses)}",
        f"2026年{i % 5 + 1}月/{rng.choice(PEOPLE)}/{rng.choice(uses)}",
    ]
    return rng.choice(kinds)


def build_workbook() -> tuple:
    from openpyxl import Workbook

    wb = Workbook()
    wb.remove(wb.active)

    rng = random.Random(20260817)  # 固定种子，可复现

    # ---- Sheet1 维护材料（inventory + flow）----
    ws = wb.create_sheet(title="维护材料")
    ws.append(WEIHU_HEADERS)
    ws.append(
        ["例", "插线板", "通用插线板 3米", "3", "个", "", "", "4", "", "",
         "2026.1.6/张伟/会议室搭建备用", "1", "2", ""]
    )
    for i in range(1, 21):
        name, spec, unit = WEIHU_MATS[i % len(WEIHU_MATS)]
        initial = rng.randint(2, 12)
        in_qty = rng.randint(0, 6) if i % 3 != 0 else 0
        out_qty = rng.randint(0, 5) if i % 4 != 0 else 0
        # 大部分行口径一致；第 6、14 行故意不一致用于勾稽差异展示
        bal = initial + in_qty - out_qty
        if i in (6, 14):
            bal = max(0, bal + rng.choice([-2, 2]))
        in_text = fmt_in_text(i, rng) if in_qty > 0 else ""
        out_text = fmt_out_text(i, rng) if out_qty > 0 else ""
        # 保持随机序列不变：照常消耗一次随机位置，仅前 5 行固定展示“213仓库”
        # （虚构存放区域，用于演示“物资种类 + 存放区域”组合筛选）
        loc = rng.choice(LOCS)
        if i in (1, 2, 3, 4, 5):
            loc = "213仓库"
        ws.append(
            [str(i), name, spec if rng.random() > 0.15 else "", bal, unit,
             loc, rng.choice(PEOPLE) if rng.random() > 0.1 else "",
             str(initial), in_text, str(in_qty) if in_qty else "",
             out_text, str(out_qty) if out_qty else "",
             str(rng.randint(1, 3)), ""]
        )

    # ---- Sheet2 备品备件（inventory + flow）----
    ws = wb.create_sheet(title="备品备件")
    ws.append(BEIPIN_HEADERS)
    ws.append(
        ["例", "整流模块", "标准整流模块 48V/40A", "6楼通信材料室", "", "3", "3", "个",
         "", "", "", "", "电源系统", "2025年度备品备件补充采购", "损坏更换", "", "8", "0", ""]
    )
    for i in range(1, 21):
        name, spec, unit = BEIPIN_MATS[i % len(BEIPIN_MATS)]
        initial = rng.randint(1, 8)
        in_qty = rng.randint(0, 4) if i % 3 != 0 else 0
        out_qty = rng.randint(0, 3) if i % 4 != 0 else 0
        bal = initial + in_qty - out_qty
        if i in (5, 12):
            bal = max(0, bal + rng.choice([-1, 1]))
        in_text = f"2025-0{i % 8 + 1}-1{i % 6 + 1} 00:00:00" if in_qty else ""
        out_text = f"2026.{i % 5 + 1}.{i % 20 + 1}/损坏更换" if out_qty else ""
        ws.append(
            [str(i), name, spec if rng.random() > 0.1 else "", rng.choice(LOCS),
             rng.choice(PEOPLE) if rng.random() > 0.15 else "", str(bal), str(initial),
             unit, in_text, str(in_qty) if in_qty else "", out_text,
             str(out_qty) if out_qty else "", rng.choice(SYSTEMS), rng.choice(PROJECTS),
             rng.choice(["损坏更换", "定期更换", "到期/损坏更换"]), rng.choice(SOURCES),
             str(rng.randint(2, 10)), str(rng.randint(0, 4)), ""]
        )

    # ---- Sheet3 应急备汛物资（inventory，无 flow）----
    ws = wb.create_sheet(title="应急备汛物资")
    ws.append(YINGJI_HEADERS)
    for i in range(1, 13):
        name, spec, unit = YINGJI_MATS[i % len(YINGJI_MATS)]
        region = "甲区" if i % 2 else "乙区"
        ws.append(
            [region, str(i), rng.choice(["防汛必备物资", "应急抢险物资", ""]),
             f"DM2026-{3000 + i}", f"SYNTH-{100000000000 + i}", name, spec, unit,
             str(rng.randint(1, 5)), str(rng.randint(1, 5)), rng.choice(LOCS),
             rng.choice(PEOPLE), "", "否", "", "", "", "", "", ""]
        )

    # ---- Sheet4 公用工器具（asset）----
    ws = wb.create_sheet(title="公用工器具")
    ws.append(GONGJU_HEADERS)
    for i in range(1, 16):
        name, spec, unit = GONGJU_MATS[i % len(GONGJU_MATS)]
        is_meter = "是" if name in ("万用表", "热成像仪", "数字万用表") else "否"
        ws.append(
            [str(i), name, spec, f"ZC2026-{500 + i}", "/" if rng.random() > 0.3 else f"ZC2026-A{i}",
             str(rng.randint(1, 4)), unit, rng.choice(LOCS), rng.choice(PEOPLE),
             is_meter, "/" if rng.random() > 0.5 else "1", "/" if rng.random() > 0.5 else "1",
             "到期/损坏更换", f"2025-{rng.randint(1, 6):02d}-{rng.randint(1, 28):02d} 00:00:00",
             rng.choice(SOURCES), "无定额", ""]
        )
    return wb, {"维护材料": 21, "备品备件": 21, "应急备汛物资": 12, "公用工器具": 15}


def main() -> int:
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "demo_data" / "samples" / "synthetic-sample.xlsx"
    wb, counts = build_workbook()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"已生成脱敏演示台账: {out}")
    print(f"各 sheet 行数（含示例行）: {counts}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
