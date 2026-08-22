# -*- coding: utf-8 -*-
"""HI-R17: 低值易耗 / 个人工器具 sheet 路由与物资 ID 隔离。"""
from __future__ import annotations

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook

from app.repositories.db import init_meta
from app.services.govern import flow_config as fc
from app.services.intake.evidence import load_stock_flow_tabular
from app.services.govern.mapping import build_domain_rows, resolve_columns


@pytest.fixture(autouse=True)
def _reload_route():
    fc._reload_ledger_route()
    yield
    fc._reload_ledger_route()


def test_ledger_route_covers_dizhi_and_geren():
    fc._reload_ledger_route()
    dizhi = fc.get_ledger_route("低值易耗")
    assert dizhi is not None
    assert dizhi["domain"] == "inventory"
    assert dizhi.get("flow") is True
    assert fc.get_ledger_route("低值易耗品") is not None
    geren = fc.get_ledger_route("个人工器具")
    assert geren is not None
    assert geren["domain"] == "asset"
    assert geren.get("flow") is False
    assert "低值易耗" in fc.ledger_sheet_names("inventory")
    assert "个人工器具" in fc.ledger_sheet_names("asset")


def _write_six_sheet_mini(path: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    dz = wb.create_sheet("低值易耗")
    dz.append(["通信部成都区域低值易耗品物资汇总表"])
    dz.append(["注意"])
    dz.append(
        [
            "序号",
            "物资名称",
            "规格型号",
            "现有库存",
            "单位",
            "存放位置",
            "入库记录",
            "入库数量",
            "出库记录",
            "出库数量",
        ]
    )
    dz.append(["例", "手持标签打印机", "WEWIN", "2", "台", "材料室", "", "", "", ""])
    dz.append(["1", "手持标签打印机", "WEWIN", "0", "台", "材料室", "2023年", "2", "", ""])

    gr = wb.create_sheet("个人工器具")
    gr.append(["", "通信部成都区域个人工器具领用记录表"])
    gr.append(
        ["序号", "类别", "型号规格", "更新周期", "", "", "", "发放人", "发放日期", "领用人", "领用数量"]
    )
    gr.append(["", "", "", "部门正副主任", "主任师", "", "", "", "", "", ""])
    gr.append(["1", "工具包", 'SATA/世达 14" 工具包', "/", "/", "/", "/", "詹雯", "2024.5.7", "吕志超", "5"])
    gr.append(["", "", "", "", "", "", "", "张停伟", "2025.3.21", "张停伟", "4"])
    gr.append(["2", "头灯", "充电防水头灯", "/", "3", "/", "3", "詹雯", "2024.5.7", "吕志超", "5"])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def test_dizhi_sheet_keeps_stock_qty(tmp_path: Path):
    fc._reload_ledger_route()
    xlsx = _write_six_sheet_mini(tmp_path / "six.xlsx")
    df = load_stock_flow_tabular(xlsx)
    dz = df[df["sheet"].astype(str) == "低值易耗"]
    assert len(dz) >= 1
    assert "stock_qty" in dz.columns
    nonempty = dz["stock_qty"].astype(str).str.strip()
    ok = (nonempty != "") & (~nonempty.str.lower().isin(["nan", "none", "null"]))
    assert int(ok.sum()) >= 1


def test_geren_sheet_routes_to_asset_items(tmp_path: Path):
    fc._reload_ledger_route()
    xlsx = _write_six_sheet_mini(tmp_path / "six.xlsx")
    df = load_stock_flow_tabular(xlsx)
    geren = df[df["sheet"].astype(str) == "个人工器具"]
    assert len(geren) >= 2
    assert "asset_name" in geren.columns
    names = set(geren["asset_name"].astype(str).str.strip())
    assert "工具包" in names
    assert "头灯" in names
    qty = pd.to_numeric(geren.loc[geren["asset_name"].astype(str) == "工具包", "asset_qty"], errors="coerce")
    assert float(qty.iloc[0]) == 9.0


def test_material_ids_isolated_by_sheet():
    init_meta()
    df = pd.DataFrame(
        [
            {"物资名称": "标签机", "现有库存": "1", "sheet": "低值易耗"},
            {"物资名称": "标签机", "现有库存": "1", "sheet": "维护材料"},
        ]
    )
    _, rows = build_domain_rows(df, domain="inventory", file_id="f1", release_id="r1", source_file="x.xlsx")
    ids = [r["material_id"] for r in rows]
    assert len(set(ids)) == 2
    assert all("低值易耗" in i or "维护材料" in i for i in ids)


REAL_305 = Path(__file__).resolve().parents[1] / "data" / "uploads" / "315653e264e9.xlsx"


@pytest.mark.skipif(not REAL_305.exists(), reason="no local 305 workbook")
def test_real_305_dizhi_and_geren_loaded():
    fc._reload_ledger_route()
    df = load_stock_flow_tabular(REAL_305)
    sheets = set(df["sheet"].astype(str))
    assert "低值易耗" in sheets
    assert "个人工器具" in sheets
    dz = df[df["sheet"].astype(str) == "低值易耗"]

    def blank(v) -> bool:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return True
        s = str(v).strip()
        return s == "" or s.lower() in {"nan", "none", "null", "-"}

    assert int(dz["stock_qty"].map(blank).sum()) < len(dz)
    geren = df[df["sheet"].astype(str) == "个人工器具"]
    assert len(geren) >= 10
    assert int(geren["asset_name"].map(blank).sum()) == 0
