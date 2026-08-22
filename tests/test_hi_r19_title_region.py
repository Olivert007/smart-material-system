# -*- coding: utf-8 -*-
"""HI-R19: infer region from sheet title rows before header."""
from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.govern import flow_config as fc
from app.services.intake.evidence import infer_sheet_region_from_title, load_stock_flow_tabular


@pytest.fixture(autouse=True)
def _reload_route():
    fc._reload_ledger_route()
    yield
    fc._reload_ledger_route()


def test_infer_sheet_region_from_title_examples():
    assert infer_sheet_region_from_title("通信部上海区域低值易耗物资汇总表") == "上海"
    assert infer_sheet_region_from_title("软件部浦东区域备品备件") == "浦东"
    assert infer_sheet_region_from_title("软件部区域 2026 年防汛") is None


def test_title_fills_region_when_column_absent(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "低值易耗"
    ws.append(["通信部上海区域低值易耗物资汇总表"])
    ws.append(["注意"])
    ws.append(["序号", "物资名称", "现有库存", "单位"])
    ws.append(["例", "示例物资", "1", "个"])
    ws.append(["1", "标签机", "2", "台"])
    xlsx = tmp_path / "dizhi.xlsx"
    wb.save(xlsx)

    df = load_stock_flow_tabular(xlsx)
    dz = df[df["sheet"].astype(str) == "低值易耗"]
    assert len(dz) >= 1
    assert "region" in dz.columns
    assert set(dz["region"].dropna().astype(str)) == {"上海"}


def test_region_column_not_overwritten(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "应急备汛物资"
    ws.append(["软件部区域 2026 年防汛"])
    ws.append(["注意"])
    ws.append(["序号", "物资名称", "区域", "现有库存", "单位"])
    ws.append(["1", "沙袋", "TD", "10", "个"])
    ws.append(["2", "水泵", "TDCD", "3", "台"])
    xlsx = tmp_path / "yj.xlsx"
    wb.save(xlsx)

    df = load_stock_flow_tabular(xlsx)
    yj = df[df["sheet"].astype(str) == "应急备汛物资"]
    assert set(yj["region"].astype(str)) == {"TD", "TDCD"}


def test_location_text_not_used_as_region(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "低值易耗"
    ws.append(["低值易耗品物资汇总表"])
    ws.append(["注意"])
    ws.append(["序号", "物资名称", "现有库存", "单位", "存放位置"])
    ws.append(["1", "标签机", "2", "台", "上海三峡大厦材料室"])
    xlsx = tmp_path / "loc.xlsx"
    wb.save(xlsx)

    df = load_stock_flow_tabular(xlsx)
    dz = df[df["sheet"].astype(str) == "低值易耗"]
    if "region" in dz.columns:
        nonempty = dz["region"].dropna().astype(str).str.strip()
        nonempty = nonempty[~nonempty.str.lower().isin({"nan", "none", ""})]
        assert "上海" not in set(nonempty)
    else:
        assert True
