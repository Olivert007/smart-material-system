# -*- coding: utf-8 -*-
"""optv3 物资规整筛选：/api/v1/materials/standardized 浏览、筛选、导出。"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from urllib.parse import unquote

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app
from app.workers import intake_worker


@pytest.fixture(autouse=True)
def _disable_worker():
    orig = intake_worker.worker.start
    intake_worker.worker.start = lambda: None
    yield
    intake_worker.worker.start = orig


@pytest.fixture()
def client():
    with TestClient(app) as c:
        yield c


def _seed():
    from app.repositories import writer_conn

    con = writer_conn()
    try:
        con.execute(
            """
            INSERT INTO dim_material
              (material_id, material_code, material_name, spec, unit, category)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            ["M-1001", "WH-001", "电力电缆", "YJV-0.6/1kV", "米", "维护材料"],
        )
        con.execute(
            """
            INSERT INTO dim_material
              (material_id, material_code, material_name, spec, unit, category)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            ["M-1002", None, "绝缘胶带", "18mm", "卷", "低值易耗品"],
        )
        con.execute(
            """
            INSERT INTO dim_material
              (material_id, material_code, material_name, spec, unit, category)
            VALUES (?, ?, ?, ?, ?, ?)
            """,
            ["M-1003", "M-1003", "内部编码物资", "X", "个", "备品备件"],
        )
        rows = [
            ("INV-1", "M-1001", "维护材料", "仓库A", 12, "rel-1", "rk-1", "台账.xlsx"),
            ("INV-2", "M-1002", "低值易耗品", "车间1区", 8, "rel-2", "rk-2", "台账.xlsx"),
            ("INV-3", "M-1003", "备品备件", "仓库A", 3, "rel-3", "rk-3", "台账.xlsx"),
            ("INV-4", "M-1001", "维护材料", "维修班组", 1, "rel-4", "rk-4", "台账.xlsx"),
        ]
        for inv_id, mid, sheet, loc, qty, rel, rk, src in rows:
            con.execute(
                """
                INSERT INTO fact_inventory
                  (inventory_id, material_id, region, category, source_file, source_sheet,
                   stock_qty, unit, location, source_release_id, row_key)
                VALUES (?, ?, '川云', ?, ?, ?, ?, '米', ?, ?, ?)
                """,
                [inv_id, mid, sheet, src, sheet, qty, loc, rel, rk],
            )
        con.execute(
            """
            INSERT INTO fact_asset
              (asset_code, asset_name, material_code, status, source_file, source_sheet,
               asset_qty, unit, location, source_release_id, row_key)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            [
                "AS-1",
                "万用表",
                "GJ-01",
                "在用",
                "台账.xlsx",
                "公用工器具",
                2,
                "台",
                "工具房",
                "rel-a1",
                "rk-a1",
            ],
        )
        con.execute(
            """
            INSERT INTO fact_inventory
              (inventory_id, material_id, region, category, source_file, source_sheet,
               stock_qty, unit, location)
            VALUES (?, ?, '川云', '维护材料', 'inject.xlsx', '维护材料', 1, '个', '=CMD')
            """,
            ["INV-INJ", "M-1002"],
        )
    finally:
        con.close()


def test_filters_returns_fixed_categories_and_locations(client):
    _seed()
    r = client.get("/api/v1/materials/standardized/filters")
    assert r.status_code == 200, r.text[:300]
    body = r.json()
    assert body["categories"] == ["维护材料", "低值易耗品", "备品备件", "公用工器具", "个人工器具"]
    assert "仓库A" in body["locations"]
    assert "车间1区" in body["locations"]
    assert "工具房" in body["locations"]


def test_default_lists_all(client):
    _seed()
    r = client.get("/api/v1/materials/standardized")
    assert r.status_code == 200, r.text[:400]
    body = r.json()
    assert body["total"] == 6
    assert len(body["items"]) == 6
    codes = {it["material_code"] for it in body["items"]}
    assert "WH-001" in codes
    assert "GJ-01" in codes
    assert "M-1001" not in codes
    assert "M-1002" not in codes
    assert None in codes  # 无正式编码 → material_code 为空，前端展示「未维护」
    tape = next(it for it in body["items"] if it["material_name"] == "绝缘胶带")
    assert tape["material_code"] is None
    assert tape["material_id"] == "M-1002"
    copied = next(it for it in body["items"] if it["material_name"] == "内部编码物资")
    assert copied["material_code"] is None
    assert copied["material_id"] == "M-1003"
    by = {x["category"]: x["count"] for x in body["summary"]["by_category"]}
    assert by["维护材料"] == 3
    assert by["低值易耗品"] == 1
    assert by["备品备件"] == 1
    assert by["公用工器具"] == 1
    assert sum(x["count"] for x in body["summary"]["by_category"]) == body["total"]


def test_filter_single_and_multi_category(client):
    _seed()
    one = client.get("/api/v1/materials/standardized", params={"categories": "维护材料"}).json()
    assert one["total"] == 3
    assert {it["category"] for it in one["items"]} == {"维护材料"}
    multi = client.get(
        "/api/v1/materials/standardized", params={"categories": "维护材料,备品备件"}
    ).json()
    assert multi["total"] == 4
    assert {it["category"] for it in multi["items"]} == {"维护材料", "备品备件"}


def test_filter_location_and_intersection(client):
    _seed()
    loc = client.get("/api/v1/materials/standardized", params={"locations": "仓库A"}).json()
    assert loc["total"] == 2
    both = client.get(
        "/api/v1/materials/standardized",
        params={"categories": "维护材料", "locations": "仓库A"},
    ).json()
    assert both["total"] == 1
    assert both["items"][0]["material_name"] == "电力电缆"
    assert both["items"][0]["location"] == "仓库A"


def test_keyword_name_and_code_not_internal_id(client):
    _seed()
    by_name = client.get("/api/v1/materials/standardized", params={"q": "电力"}).json()
    assert by_name["total"] == 2
    assert {it["material_name"] for it in by_name["items"]} == {"电力电缆"}
    by_code = client.get("/api/v1/materials/standardized", params={"q": "WH-001"}).json()
    assert by_code["total"] == 2
    by_internal = client.get("/api/v1/materials/standardized", params={"q": "M-1002"}).json()
    assert by_internal["total"] == 0


def test_unknown_category_returns_empty_intersection(client):
    _seed()
    r = client.get("/api/v1/materials/standardized", params={"categories": "不存在,维护材料"})
    assert r.status_code == 200
    body = r.json()
    assert body["filters"]["categories"] == ["不存在", "维护材料"]
    assert body["total"] == 3
    assert {it["category"] for it in body["items"]} == {"维护材料"}


def _xlsx_rows(content: bytes) -> list[tuple]:
    wb = load_workbook(BytesIO(content))
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def test_export_follows_filter_and_rejects_empty(client):
    _seed()
    ok = client.get(
        "/api/v1/materials/standardized/export",
        params={"categories": "备品备件", "locations": "仓库A"},
    )
    assert ok.status_code == 200, ok.text[:300]
    assert "spreadsheetml" in ok.headers.get("content-type", "")
    disp = unquote(ok.headers.get("content-disposition") or "")
    assert f"物资台账_筛选结果_{datetime.now().strftime('%Y%m%d')}" in disp
    rows = _xlsx_rows(ok.content)
    header = [str(x) for x in rows[0]]
    assert header[0] == "物资编码"
    assert "物资名称" in header and "存放区域" in header
    assert "material_id" not in header and "row_key" not in header
    assert len(rows) == 2
    values = " ".join(str(c) for c in rows[1])
    assert "内部编码物资" in values
    assert "未维护" in values
    assert "M-1003" not in values

    empty = client.get(
        "/api/v1/materials/standardized/export",
        params={"categories": "个人工器具"},
    )
    assert empty.status_code == 400
    assert empty.json()["code"] == "EMPTY_EXPORT"


def test_export_formula_injection_and_no_internal_id(client):
    _seed()
    r = client.get("/api/v1/materials/standardized/export", params={"locations": "=CMD"})
    assert r.status_code == 200
    rows = _xlsx_rows(r.content)
    header = [str(x) for x in rows[0]]
    assert "M-1002" not in header
    loc_idx = header.index("存放区域")
    assert any(str(row[loc_idx] or "").startswith("'=") for row in rows[1:])


def test_sort_whitelist_ignores_injection(client):
    _seed()
    injected = client.get(
        "/api/v1/materials/standardized",
        params={"sort_by": "stock_qty;drop table", "sort_order": "desc"},
    )
    assert injected.status_code == 200
    assert injected.json()["total"] == 6
    desc = client.get(
        "/api/v1/materials/standardized",
        params={"sort_by": "stock_qty", "sort_order": "desc"},
    ).json()
    qtys = [it["stock_qty"] for it in desc["items"]]
    assert qtys == sorted(qtys, reverse=True)
    assert qtys[0] == 12


def test_empty_db_lists_zero_and_rejects_export(client):
    listed = client.get("/api/v1/materials/standardized").json()
    assert listed["total"] == 0
    assert listed["items"] == []
    filt = client.get("/api/v1/materials/standardized/filters").json()
    assert filt["categories"][0] == "维护材料"
    assert filt["locations"] == []
    exp = client.get("/api/v1/materials/standardized/export")
    assert exp.status_code == 400
    assert exp.json()["code"] == "EMPTY_EXPORT"


def test_pagination_and_offset_past_end(client):
    _seed()
    p1 = client.get("/api/v1/materials/standardized", params={"limit": 2, "offset": 0, "sort_by": "stock_qty", "sort_order": "desc"}).json()
    p2 = client.get("/api/v1/materials/standardized", params={"limit": 2, "offset": 2, "sort_by": "stock_qty", "sort_order": "desc"}).json()
    assert p1["total"] == 6
    assert len(p1["items"]) == 2
    assert len(p2["items"]) == 2
    keys1 = {(it["material_name"], it["location"]) for it in p1["items"]}
    keys2 = {(it["material_name"], it["location"]) for it in p2["items"]}
    assert keys1.isdisjoint(keys2)
    past = client.get("/api/v1/materials/standardized", params={"limit": 20, "offset": 100}).json()
    assert past["total"] == 6
    assert past["items"] == []


def test_keyword_wildcard_and_sql_chars_are_literal(client):
    _seed()
    wild = client.get("/api/v1/materials/standardized", params={"q": "%"}).json()
    assert wild["total"] == 0
    injected = client.get("/api/v1/materials/standardized", params={"q": "电力' OR 1=1 --"}).json()
    assert injected["total"] == 0
    underscore = client.get("/api/v1/materials/standardized", params={"q": "电力_缆"}).json()
    assert underscore["total"] == 0
    exact = client.get("/api/v1/materials/standardized", params={"q": "电力电缆"}).json()
    assert exact["total"] == 2


def test_locations_or_and_export_count_matches(client):
    _seed()
    listed = client.get(
        "/api/v1/materials/standardized",
        params={"locations": "仓库A,车间1区"},
    ).json()
    assert listed["total"] == 3
    assert {it["location"] for it in listed["items"]} == {"仓库A", "车间1区"}
    exp = client.get(
        "/api/v1/materials/standardized/export",
        params={"locations": "仓库A,车间1区"},
    )
    assert exp.status_code == 200
    rows = _xlsx_rows(exp.content)
    assert len(rows) - 1 == listed["total"]


def test_orphan_inventory_and_asset_without_code(client):
    from app.repositories import writer_conn

    con = writer_conn()
    try:
        con.execute(
            """
            INSERT INTO fact_inventory
              (inventory_id, material_id, region, category, source_file, source_sheet,
               stock_qty, unit, location)
            VALUES ('INV-ORPH', 'M-GONE', '川云', '维护材料', 'x.xlsx', '维护材料', 4, '个', '工具房')
            """
        )
        con.execute(
            """
            INSERT INTO fact_asset
              (asset_code, asset_name, status, source_file, source_sheet,
               asset_qty, unit, location)
            VALUES ('AS-NOCODE', '绝缘手套', '在用', 'x.xlsx', '个人工器具', 1, '双', '维修班组')
            """
        )
    finally:
        con.close()
    body = client.get("/api/v1/materials/standardized").json()
    assert body["total"] == 2
    orphan = next(it for it in body["items"] if it["location"] == "工具房")
    assert orphan["material_code"] is None
    assert orphan["material_name"] == ""
    assert orphan["material_id"] == "M-GONE"
    glove = next(it for it in body["items"] if it["material_name"] == "绝缘手套")
    assert glove["material_code"] is None
    assert glove["category"] == "个人工器具"
    by = {x["category"]: x["count"] for x in body["summary"]["by_category"]}
    assert by["维护材料"] + by["个人工器具"] == 2
    assert sum(x["count"] for x in body["summary"]["by_category"]) == body["total"]


def test_e2e_browse_filter_export_flow(client):
    _seed()
    start = client.get("/api/v1/materials/standardized").json()
    assert start["total"] == 6
    filt = client.get("/api/v1/materials/standardized/filters").json()
    assert "仓库A" in filt["locations"]
    narrowed = client.get(
        "/api/v1/materials/standardized",
        params={"categories": "维护材料", "q": "电缆"},
    ).json()
    assert narrowed["total"] == 2
    assert all(it["category"] == "维护材料" for it in narrowed["items"])
    assert all("电缆" in it["material_name"] for it in narrowed["items"])
    assert sum(x["count"] for x in narrowed["summary"]["by_category"]) == narrowed["total"]
    exp = client.get(
        "/api/v1/materials/standardized/export",
        params={"categories": "维护材料", "q": "电缆"},
    )
    assert exp.status_code == 200
    rows = _xlsx_rows(exp.content)
    header = [str(x) for x in rows[0]]
    assert "物资编码" in header
    assert "source_release_id" not in header
    assert len(rows) - 1 == 2
    empty = client.get(
        "/api/v1/materials/standardized/export",
        params={"categories": "维护材料", "q": "不存在的物资"},
    )
    assert empty.status_code == 400
    assert empty.json()["code"] == "EMPTY_EXPORT"
