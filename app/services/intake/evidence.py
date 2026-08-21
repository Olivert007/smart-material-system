# -*- coding: utf-8 -*-
"""Multi-format file → raw evidence parquet."""
from __future__ import annotations

import hashlib
import json
from pathlib import Path

import openpyxl
import pandas as pd

from app import config

# xlsx stays openpyxl read_only; other spreadsheet formats use calamine (docs/03 §2.3).
_CALAMINE_EXTS = {"xls", "xlsb", "xlsm", "ods"}
_OPENPYXL_EXTS = {"xlsx"}
_OCR_EXTS = {"pdf", "png", "jpg", "jpeg", "webp", "bmp", "tif", "tiff"}
_SUPPORTED = {"xlsx", "xls", "xlsb", "xlsm", "ods", "csv", "json"} | _OCR_EXTS


def col_letter(idx: int) -> str:
    s = ""
    idx += 1
    while idx:
        idx, r = divmod(idx - 1, 26)
        s = chr(65 + r) + s
    return s


def sha256_file(path: Path, chunk: int = 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            b = f.read(chunk)
            if not b:
                break
            h.update(b)
    return h.hexdigest()


def _df_sheet_to_evidence(
    sheet_df: pd.DataFrame,
    *,
    file_id: str,
    sheet: str,
    max_evidence: int,
) -> list[dict]:
    rows: list[dict] = []
    n = 0
    for _, row in sheet_df.iterrows():
        n += 1
        if n > max_evidence:
            rows.append(
                {
                    "file_id": file_id,
                    "sheet": sheet,
                    "row": -1,
                    "col": "*",
                    "raw_value": f"[sparse sheet truncated at {max_evidence} rows]",
                    "value_type": "marker",
                }
            )
            break
        for c, v in enumerate(row.tolist(), 1):
            if v is None or (isinstance(v, float) and v != v):
                continue
            s = str(v).strip()
            if s == "" or s.lower() == "nan":
                continue
            rows.append(
                {
                    "file_id": file_id,
                    "sheet": sheet,
                    "row": n,
                    "col": col_letter(c - 1),
                    "raw_value": s,
                    "value_type": type(v).__name__,
                }
            )
    return rows


def _load_xlsx(path: Path, file_id: str, max_evidence: int) -> tuple[pd.DataFrame, int]:
    rows: list[dict] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    n_sheets = len(wb.sheetnames)
    for ws in wb.worksheets:
        n = 0
        for r in ws.iter_rows(values_only=True):
            n += 1
            if n > max_evidence:
                break
            for c, v in enumerate(r, 1):
                if v is None or (isinstance(v, float) and v != v):
                    continue
                rows.append(
                    {
                        "file_id": file_id,
                        "sheet": ws.title,
                        "row": n,
                        "col": col_letter(c - 1),
                        "raw_value": str(v),
                        "value_type": type(v).__name__,
                    }
                )
        if n > max_evidence:
            rows.append(
                {
                    "file_id": file_id,
                    "sheet": ws.title,
                    "row": -1,
                    "col": "*",
                    "raw_value": f"[sparse sheet truncated at {max_evidence} rows]",
                    "value_type": "marker",
                }
            )
    wb.close()
    return pd.DataFrame(rows), n_sheets


def _load_calamine(path: Path, file_id: str, max_evidence: int) -> tuple[pd.DataFrame, int]:
    """xls / xlsb / xlsm / ods via python-calamine (pandas engine=calamine)."""
    try:
        book = pd.read_excel(
            path, sheet_name=None, header=None, dtype=object, engine="calamine"
        )
    except Exception as e:
        raise ValueError(f"calamine failed for {path.name}: {e}") from e
    if not book:
        return pd.DataFrame(
            columns=["file_id", "sheet", "row", "col", "raw_value", "value_type"]
        ), 0
    rows: list[dict] = []
    for name, sdf in book.items():
        if sdf is None or sdf.empty:
            continue
        rows.extend(
            _df_sheet_to_evidence(
                sdf, file_id=file_id, sheet=str(name), max_evidence=max_evidence
            )
        )
    return pd.DataFrame(rows), len(book)


def _expand_xlsx_merged_cells(path: Path) -> dict[str, pd.DataFrame]:
    """Read xlsx and copy merged-cell parent values into empty children.

    pandas/openpyxl leave child cells as NA, so 305 维护材料「南孚电池」等
    合并名称会被当成空 material_name 拦截（HI-R16）。
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    out: dict[str, pd.DataFrame] = {}
    try:
        for ws in wb.worksheets:
            ranges = list(ws.merged_cells.ranges)
            fills: list[tuple[int, int, int, int, object]] = []
            for rng in ranges:
                val = ws.cell(rng.min_row, rng.min_col).value
                fills.append((rng.min_row, rng.max_row, rng.min_col, rng.max_col, val))
            for rng in ranges:
                try:
                    ws.unmerge_cells(str(rng))
                except Exception:
                    continue
            for min_r, max_r, min_c, max_c, val in fills:
                if val is None or str(val).strip() == "":
                    continue
                # 只展开纵向合并（名称列 B125:B128）；横向标题 A1:N1 填满会污染表头探测。
                if max_r <= min_r:
                    continue
                for r in range(min_r, max_r + 1):
                    for c in range(min_c, max_c + 1):
                        cell = ws.cell(r, c)
                        if cell.value is None or str(cell.value).strip() == "":
                            cell.value = val
            rows = [list(row) for row in ws.iter_rows(values_only=True)]
            out[ws.title] = pd.DataFrame(rows)
    finally:
        wb.close()
    return out


def _apply_excel_read_kwargs(df: pd.DataFrame, **kwargs) -> pd.DataFrame:
    header = kwargs.get("header", 0)
    if df is None or df.empty:
        out = df.copy() if df is not None else pd.DataFrame()
    elif header is None:
        out = df.copy()
        out.columns = list(range(len(out.columns)))
    else:
        hdr = int(header)
        if hdr >= len(df):
            out = df.copy()
        else:
            cols = [
                str(x).strip() if x is not None else f"Unnamed: {i}"
                for i, x in enumerate(df.iloc[hdr].tolist())
            ]
            out = df.iloc[hdr + 1 :].copy()
            out.columns = cols
            out.reset_index(drop=True, inplace=True)
    dtype = kwargs.get("dtype")
    if dtype is str and not out.empty:
        def _to_str(v):
            if v is None or (isinstance(v, float) and pd.isna(v)):
                return None
            return str(v)

        out = out.apply(lambda col: col.map(_to_str))
    return out


def _read_excel_best(path: Path, **kwargs):
    """Prefer calamine for non-xlsx; xlsx expands merged cells then returns frames."""
    ext = path.suffix.lstrip(".").lower()
    if ext in _CALAMINE_EXTS:
        kwargs.setdefault("engine", "calamine")
        return pd.read_excel(path, **kwargs)
    if ext in _OPENPYXL_EXTS:
        book = _expand_xlsx_merged_cells(path)
        sheet_name = kwargs.get("sheet_name", 0)
        if sheet_name is None:
            return {name: _apply_excel_read_kwargs(frame, **kwargs) for name, frame in book.items()}
        if isinstance(sheet_name, int):
            names = list(book.keys())
            if not names:
                return pd.DataFrame()
            name = names[sheet_name]
            return _apply_excel_read_kwargs(book[name], **kwargs)
        if sheet_name in book:
            return _apply_excel_read_kwargs(book[sheet_name], **kwargs)
        return pd.DataFrame()
    return pd.read_excel(path, **kwargs)


def _load_csv(path: Path, file_id: str) -> tuple[pd.DataFrame, int]:
    # header=None keeps the physical header row in evidence (row 1), so that
    # profile/header detection and map_pending see real column headers instead
    # of the first data row's values. Matches XLSX evidence row convention.
    try:
        df = pd.read_csv(path, dtype=str, header=None, on_bad_lines="skip")
    except UnicodeDecodeError:
        df = pd.read_csv(path, dtype=str, header=None, encoding="gbk", on_bad_lines="skip")
    rows = []
    for i, (_, row) in enumerate(df.iterrows(), 1):
        for c, v in enumerate(row, 1):
            if pd.isna(v) or str(v).strip() == "":
                continue
            rows.append(
                {
                    "file_id": file_id,
                    "sheet": "Sheet1",
                    "row": i,
                    "col": col_letter(c - 1),
                    "raw_value": str(v),
                    "value_type": "str",
                }
            )
    return pd.DataFrame(rows), 1


def _load_json(path: Path, file_id: str) -> tuple[pd.DataFrame, int]:
    with path.open(encoding="utf-8") as f:
        data = json.load(f)
    if isinstance(data, dict):
        data = [data]
    rows = []
    for i, rec in enumerate(data, 1):
        if isinstance(rec, dict):
            for c, (_k, v) in enumerate(rec.items(), 1):
                if v is None:
                    continue
                rows.append(
                    {
                        "file_id": file_id,
                        "sheet": "json",
                        "row": i,
                        "col": col_letter(c - 1),
                        "raw_value": str(v),
                        "value_type": type(v).__name__,
                    }
                )
    return pd.DataFrame(rows), 1


def _header_alias_hits(cols: list[str], domain: str = "inventory") -> int:
    from app.services.mapping import ALIASES, _canon_header, _norm

    alias_norms = {_norm(n) for names in ALIASES.get(domain, {}).values() for n in names}
    alias_canons = {_canon_header(n) for names in ALIASES.get(domain, {}).values() for n in names}
    hits = 0
    for c in cols:
        cn = _norm(c)
        cc = _canon_header(c)
        if cn in alias_norms or cc in alias_canons:
            hits += 1
            continue
        if any(cn.startswith(a) or (len(a) >= 2 and a in cc) for a in alias_canons):
            hits += 1
    return hits


def normalize_tabular(df: pd.DataFrame, *, domain: str = "inventory", max_probe: int = 8) -> pd.DataFrame:
    """Promote a real header row when Excel title/merged rows pollute columns."""
    if df is None or df.empty:
        return df
    best = df
    best_score = _header_alias_hits([str(c) for c in df.columns], domain)
    unnamed = sum(1 for c in df.columns if str(c).startswith("Unnamed") or str(c).lower() == "nan")
    if best_score >= 3 and unnamed <= max(1, len(df.columns) // 4):
        return best

    raw = df.copy()
    probe = pd.concat(
        [pd.DataFrame([[str(c) for c in raw.columns]]), raw.astype(str)],
        ignore_index=True,
    )
    limit = min(max_probe, len(probe) - 1)
    for i in range(limit):
        header = [str(x).strip() for x in probe.iloc[i].tolist()]
        if sum(1 for h in header if h and h.lower() not in {"nan", "none", ""}) < 3:
            continue
        body = probe.iloc[i + 1 :].copy()
        body.columns = header
        body = body.loc[:, [c for c in body.columns if str(c).strip() and str(c).lower() != "nan"]]
        score = _header_alias_hits([str(c) for c in body.columns], domain)
        if score > best_score:
            best_score = score
            best = body.reset_index(drop=True)
    return best


def normalize_tabular_best(df: pd.DataFrame) -> tuple[pd.DataFrame, str, int]:
    """Pick inventory/asset/demand/stock_flow header projection with highest alias hits."""
    best_df = df
    best_domain = "inventory"
    best_score = -1
    for domain in ("inventory", "asset", "demand", "stock_flow"):
        cand = normalize_tabular(df, domain=domain)
        score = _header_alias_hits([str(c) for c in cand.columns], domain)
        if score > best_score:
            best_df, best_domain, best_score = cand, domain, score
    return best_df, best_domain, best_score


def _cell_blank(v: object) -> bool:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return True
    s = str(v).strip()
    return s == "" or s.lower() in {"nan", "none", "null", "-"}


def _adapt_personal_tools_sheet(raw: pd.DataFrame) -> pd.DataFrame:
    """个人工器具领用记录 → 类别+型号一行，数量=领用合计。"""
    header_idx = None
    cat_col = spec_col = qty_col = None
    user_col = giver_col = date_col = seq_col = None
    for i in range(min(12, len(raw))):
        vals = ["" if _cell_blank(x) else str(x).strip() for x in raw.iloc[i].tolist()]
        if "类别" in vals and any("领用数量" in v for v in vals):
            header_idx = i
            for j, v in enumerate(vals):
                if v == "序号":
                    seq_col = j
                elif v == "类别":
                    cat_col = j
                elif v in {"型号规格", "规格型号"}:
                    spec_col = j
                elif "领用数量" in v:
                    qty_col = j
                elif v == "领用人":
                    user_col = j
                elif v == "发放人":
                    giver_col = j
                elif v in {"发放日期", "领用日期"}:
                    date_col = j
            break
    if header_idx is None or cat_col is None or qty_col is None:
        return pd.DataFrame()
    body = raw.iloc[header_idx + 1 :].copy()
    if seq_col is not None:
        seq = body.iloc[:, seq_col].map(lambda v: "" if _cell_blank(v) else str(v).strip())
        body = body.loc[~seq.isin({"例", "示例", "example"})]
    if body.empty:
        return pd.DataFrame()

    def series(idx: int | None) -> pd.Series:
        if idx is None:
            return pd.Series([None] * len(body), index=body.index)
        return body.iloc[:, idx]

    cat = series(cat_col).map(lambda v: None if _cell_blank(v) else str(v).strip())
    spec = series(spec_col).map(lambda v: None if _cell_blank(v) else str(v).strip())
    cat = cat.ffill()
    spec = spec.ffill()
    qty = pd.to_numeric(series(qty_col), errors="coerce")
    work = pd.DataFrame(
        {
            "类别": cat,
            "型号规格": spec,
            "数量": qty,
            "使用人": series(user_col).map(lambda v: None if _cell_blank(v) else str(v).strip()),
            "发放人": series(giver_col).map(lambda v: None if _cell_blank(v) else str(v).strip()),
            "购买日期": series(date_col).map(lambda v: None if _cell_blank(v) else str(v).strip()),
        },
        index=body.index,
    )
    work = work[work["数量"].notna() & work["类别"].notna()]
    if work.empty:
        return pd.DataFrame()
    grouped = work.groupby(["类别", "型号规格"], dropna=False, as_index=False).agg(
        数量=("数量", "sum"),
        使用人=("使用人", "first"),
        发放人=("发放人", "first"),
        购买日期=("购买日期", "first"),
    )
    return pd.DataFrame(
        {
            "资产名称": grouped["类别"],
            "型号规格": grouped["型号规格"],
            "数量": grouped["数量"],
            "使用人": grouped["使用人"],
            "管理者": grouped["发放人"],
            "购买日期": grouped["购买日期"],
        }
    )


def _ledger_keep_fields() -> tuple[str, ...]:
    """T3.2: 标准字段并集（inventory/asset/stock_flow）+ sheet 标记。"""
    from app.services.mapping import ALIASES

    fields: set[str] = set()
    for dom in ("inventory", "asset", "stock_flow"):
        fields.update(ALIASES.get(dom, {}).keys())
    fields.discard("source_sheet")
    return tuple(sorted(fields)) + ("sheet",)


def load_stock_flow_tabular(path: Path) -> pd.DataFrame:
    """Load sheets: 305B/ZW flow ledgers AND ledger-route sheets (T3.1); tag `sheet` column.

    - flow=true 路由 sheet（维护材料/备品备件/低值易耗）→ stock_flow 提取并合并库存列
    - flow=false 路由 sheet（公用工器具/个人工器具→asset、应急备汛→inventory）→ 域投影
    - 个人工器具：领用记录 adapter，按类别+型号合计领用数量
    - 未命中路由 → 旧 flow 判定（无流水列则跳过，保持既有文件行为不变）
    """
    from app.services.govern.flow_config import get_ledger_route
    from app.services.mapping import resolve_columns

    try:
        book = _read_excel_best(path, sheet_name=None, dtype=str, header=None)
    except Exception:
        return pd.DataFrame()

    frames: list[pd.DataFrame] = []
    keep_fields = _ledger_keep_fields()
    for name, raw in book.items():
        if raw is None or raw.empty:
            continue
        route = get_ledger_route(str(name))
        if route is not None:
            domain = "stock_flow" if route.get("flow") else str(route.get("domain") or "stock_flow")
        else:
            domain = "stock_flow"
        if route is not None and str(route.get("sheet") or "") == "个人工器具":
            df = _adapt_personal_tools_sheet(raw)
            domain = "asset"
            if df is None or df.empty:
                continue
        else:
            df = normalize_tabular(raw, domain=domain, max_probe=12)
        mapping = resolve_columns(df, domain, source_sheet=str(name))
        if route is not None and domain == "stock_flow":
            # T10.2 修复：flow sheet 仅投影流水列 → 库存快照列（stock_qty/opening_qty/
            # location/custodian/…）全空；合并 inventory 映射补回，stock_flow 映射优先。
            inv_map = resolve_columns(df, "inventory", source_sheet=str(name))
            mapping = {**inv_map, **mapping}
        elif route is None:
            # T11: 非台账路由文件（如维护材料采购订单）若只按 stock_flow 投影，
            # 「数量/规格/位置/区域」等库存列会被丢弃 → 质量预检 REQUIRED_UNMAPPED
            # 阻塞无法发布。合并 inventory 映射补回库存快照列，stock_flow 映射优先。
            inv_map = resolve_columns(df, "inventory", source_sheet=str(name))
            mapping = {**inv_map, **mapping}
        if route is None:
            if "flow_in_text" not in mapping and "flow_out_text" not in mapping:
                continue
            if "material_name" not in mapping and "material_code" not in mapping:
                continue
        drop_mask = None
        for c in df.columns:
            if str(c).strip() in {"序号", "seq", "No", "no"}:
                col = df[c]
                if isinstance(col, pd.DataFrame):
                    col = col.iloc[:, 0]
                ser = col.astype(str).str.strip()
                drop_mask = ser.isin({"例", "示例", "example"})
                break
        if drop_mask is None:
            c0 = df.columns[0]
            ser = df.iloc[:, 0] if isinstance(df[c0], pd.DataFrame) else df[c0]
            ser = ser.astype(str).str.strip()
            drop_mask = ser.isin({"例", "示例", "example"})
        df = df.loc[~drop_mask].reset_index(drop=True)
        if len(df) == 0:
            continue
        std = pd.DataFrame()
        for field, col in mapping.items():
            if field in keep_fields and col in df.columns:
                val = df[col]
                if isinstance(val, pd.DataFrame):
                    val = val.iloc[:, 0]
                std[field] = val
        std["sheet"] = str(name)
        id_cols = [c for c in ("material_name", "material_code", "asset_name", "asset_code") if c in std.columns]
        if id_cols:
            keep_id = ~std[id_cols].apply(lambda r: all(_cell_blank(x) for x in r), axis=1)
            std = std.loc[keep_id].reset_index(drop=True)
        if len(std) == 0:
            continue
        frames.append(std)

    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True)


def load_to_evidence(path: Path, file_id: str) -> tuple[pd.DataFrame, str, int, pd.DataFrame | None]:
    """Return (cell_evidence, format, n_sheets, tabular_df|None)."""
    ext = path.suffix.lstrip(".").lower()
    if ext not in _SUPPORTED:
        raise ValueError(f"unsupported format .{ext}; supported: {sorted(_SUPPORTED)}")
    tabular: pd.DataFrame | None = None
    if ext in _OPENPYXL_EXTS:
        df, n_sheets = _load_xlsx(path, file_id, config.SPARSE_EVIDENCE_ROWS)
        flow_tab = load_stock_flow_tabular(path)
        if flow_tab is not None and len(flow_tab) > 0:
            tabular = flow_tab
        else:
            try:
                tabular, _domain, _score = normalize_tabular_best(
                    _read_excel_best(path, sheet_name=0, dtype=str, header=None)
                )
            except Exception:
                try:
                    tabular, _domain, _score = normalize_tabular_best(
                        _read_excel_best(path, sheet_name=0, dtype=str)
                    )
                except Exception:
                    tabular = None
    elif ext in _CALAMINE_EXTS:
        df, n_sheets = _load_calamine(path, file_id, config.SPARSE_EVIDENCE_ROWS)
        flow_tab = load_stock_flow_tabular(path)
        if flow_tab is not None and len(flow_tab) > 0:
            tabular = flow_tab
        else:
            try:
                tabular, _domain, _score = normalize_tabular_best(
                    _read_excel_best(path, sheet_name=0, dtype=str, header=None)
                )
            except Exception:
                tabular = None
    elif ext == "csv":
        df, n_sheets = _load_csv(path, file_id)
        try:
            tabular = pd.read_csv(path, dtype=str, on_bad_lines="skip")
        except UnicodeDecodeError:
            tabular = pd.read_csv(path, dtype=str, encoding="gbk", on_bad_lines="skip")
        tabular, _domain, _score = normalize_tabular_best(tabular)
    elif ext in _OCR_EXTS:
        from app.services.ocr_evidence import load_ocr_evidence

        df, n_sheets, tabular, _meta = load_ocr_evidence(path, file_id)
        if tabular is not None and len(tabular) > 0:
            try:
                tabular, _domain, _score = normalize_tabular_best(tabular)
            except Exception:
                pass
    else:
        df, n_sheets = _load_json(path, file_id)
        with path.open(encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, dict):
            data = [data]
        if data and isinstance(data[0], dict):
            tabular = pd.DataFrame(data)
    return df, ext, n_sheets, tabular


def save_evidence(df: pd.DataFrame, file_id: str, tabular: pd.DataFrame | None = None) -> Path:
    out = config.RAW / f"{file_id}.parquet"
    df.to_parquet(out, index=False)
    if tabular is not None and len(tabular) > 0:
        tab_path = config.RAW / f"{file_id}.tabular.parquet"
        tabular.to_parquet(tab_path, index=False)
    return out


def evidence_path(file_id: str) -> Path:
    return config.RAW / f"{file_id}.parquet"


def tabular_path(file_id: str) -> Path:
    return config.RAW / f"{file_id}.tabular.parquet"
